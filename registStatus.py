from email.header import decode_header
from bs4 import BeautifulSoup as bs
import xlrd
import openpyxl
import datetime
import time
from dotenv import load_dotenv
import os
from imapclient import IMAPClient
from mailparser import parse_from_bytes
import re
from status import status_

load_dotenv()


def get_all_drs():

    # Configurações do servidor IMAP do Outlook
    outlook_server = "mail.aquinos.org"
    outlook_username = "aqacademy@aquinos.pt"
    outlook_password = "P@ssw0rd01"


    # Conectando ao servidor IMAP do Outlook
    with IMAPClient(outlook_server) as client:
        client.login(outlook_username, outlook_password)
        # Selecionando a caixa de entrada
        client.select_folder("INBOX")


        # Buscando os IDs dos emails na caixa de entrada
        # Você pode ajustar os critérios de busca conforme necessário
        messages = client.search("ALL")

        #reverse all mesages
        messages.sort(reverse=True)
        # Iterando sobre os IDs dos emails
        # get all drs non duplicated 
        all_drs = []
        controler = []

        count = 0
        for msg_id in messages:
            if count == 15:
                break
        # Obtendo o corpo do email
            raw_message = client.fetch([msg_id], ["RFC822"])[msg_id][b"RFC822"]
            parsed_message = parse_from_bytes(raw_message)


            # Extraindo informações do email
            sender = parsed_message.from_
            date_sent = parsed_message.date


            # Tentando extrair diferentes partes do corpo do e-mail
            body_parts = []


            # Texto simples
            if parsed_message.text_plain:
                body_parts.append(parsed_message.text_plain)


            # HTML
            if parsed_message.text_html:
                body_parts.append(parsed_message.text_html)


            # Outras partes
            if parsed_message.text_other:
                body_parts.append(parsed_message.text_other)


            # Exibindo informações do email
            #print(f"De: {sender}")
            #print(f"Data de Envio: {date_sent}")


            # Exibindo todas as partes do corpo do e-mail
            #re pattern STATUS: DRS [1-9]\d{3} (?:APROVADA|FINALIZADA|ANULADA|RECUSADA)
            pattern = re.compile(r'STATUS: DRS [1-9]\d{3} (?:APROVADA|FINALIZADA|ANULADA)(?: \d{2}/\d{2}/\d{4})?')


            if body_parts:
                #print("Corpo(s) do Email:")
                
                for part in body_parts:
                    for p in part:
                        matches = pattern.findall(p.upper())
                        for m in matches:
                            #print(m)
                            # object
                            if m not in controler:
                                
                                #requester
                                requester = sender[0][0]
                                #date ( check if date has given )
                                if len(m.split(' ')) > 4:
                                    date = m.split(' ')[-1]
                                else:
                                    date = date_sent.strftime('%d/%m/%Y')

                                #drs number
                                drs_number = m.split(' ')[2]

                                #status
                                status = m.split(' ')[3]
                                
                                to_add = {
                                    'requester':requester,
                                    'date':date,
                                    'drsnumber':drs_number,
                                    'status':status
                                }
                                all_drs.append(to_add)
                            controler.append(m)
            
                count += 1
            else:
                print("Nenhuma parte do corpo do e-mail encontrada.")
    return all_drs





def getDrsStatus():
    print('     -> Run DRS Status...')

    #Get drs data
    drs_data = get_all_drs()

    #drs = parseData(data)

    path_db = os.environ.get('DIRDATABASE')
    
    #open db
    db_file = xlrd.open_workbook(path_db).sheet_by_index(0)

    #loop to drs_data
    for _ in drs_data:
        drs_number = _['drsnumber']     
        requester = _['requester']
        status = _['status']
        date = _['date']
        
        #check if drs in db and the status
        for x in range(db_file.nrows):
            drs_n = db_file.cell_value(x, 0)
            if(type(drs_n) != str):
                if int(drs_n) == int(drs_number):
                    if status.upper() == 'FINALIZADA':
                        #check if this status are empty
                        isStatusEmpty = db_file.cell_value(x, 27)
                        if len(isStatusEmpty) == 0:
                            #add status to DB
                            file_ = openpyxl.load_workbook(path_db)
                            dbsheet = file_.active
                            dbsheet[f'AB{x + 1}'] = requester
                            dbsheet[f'AC{x + 1}'] = date
                            file_.save(path_db)
                            file_.close()
                            print(f'DRS {drs_n} add to db with status: {status.upper()}')


                    if status.upper() == 'APROVADA':
                        #check if this status are empty

                        isStatusEmpty = db_file.cell_value(x, 29)
                        if len(isStatusEmpty) == 0:
                            #add status to DB
                            file_ = openpyxl.load_workbook(path_db)
                            dbsheet = file_.active
                            dbsheet[f'AD{x + 1}'] = requester
                            dbsheet[f'AF{x + 1}'] = date
                            file_.save(path_db)
                            file_.close()
                            print(f'DRS {drs_n} add to db with status: {status.upper()}')

                    if status.upper() == 'RECUSADA':
                        #check if this status are empty

                        isStatusEmpty = db_file.cell_value(x, 29)
                        isStatusEmpty_2 = db_file.cell_value(x, 30)
                        if len(isStatusEmpty) == 0 and len(isStatusEmpty_2) == 0:
                            #add status to DB
                            file_ = openpyxl.load_workbook(path_db)
                            dbsheet = file_.active
                            dbsheet[f'AE{x + 1}'] = requester
                            dbsheet[f'AF{x + 1}'] = date
                            file_.save(path_db)
                            file_.close()
                            print(f'DRS {drs_n} add to db with status: {status.upper()}')
                    
                    if status.upper() == 'ANULADA':
                        #check if this status are empty

                        isStatusEmpty = db_file.cell_value(x, 32)
                        if len(isStatusEmpty) == 0:
                            #add status to DB
                            file_ = openpyxl.load_workbook(path_db)
                            dbsheet = file_.active
                            dbsheet[f'AG{x + 1}'] = requester
                            dbsheet[f'AH{x + 1}'] = date
                            file_.save(path_db)
                            file_.close()
                            print(f'DRS {drs_n} add to db with status: {status.upper()}')

            

def registStatus():
    getDrsStatus()

