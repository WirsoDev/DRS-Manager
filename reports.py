#make reports by email - use AQsmtpAPI
from openpyxl import load_workbook
import datetime
import xlrd
import os
from dotenv import load_dotenv

load_dotenv()

class REPORTS:
    def __init__(self, global_=False):
        #Get all data from DRS DB

        date = datetime.datetime.now()

        db_file = os.environ.get('DIRDATABASE')
        file = load_workbook(filename=db_file)
        self.sheet = file['Folha1']
        self.global_ = global_
        self.filter_year = date.year

        max = self.sheet.max_row

        self.current_mounth_drs = {}

        x = 3
        while True:
            value = self.sheet[f'AA{x}'].value
            drs_number = self.sheet[f'A{x}'].value
            if value == None:
                break
            str_value = str(value)
            recived_month = str_value.split('-')[1]
            recived_year = str_value.split('-')[0]
            # remove left 0
            if recived_month[0] == '0':
                recived_month = recived_month.replace('0', '')

            
            if date.month == 1:
                past_month = 12
                self.year = date.year - 1
            else:
                past_month = date.month - 1
                self.year = date.year
            
            
            drs_name = f"{self.sheet[f'A{x}'].value}_{str(self.sheet[f'B{x}'].value)}"


            #controller if is global data or pass month
            if self.global_:
                if int(recived_year) == int(self.filter_year):
                    new_dic = {drs_name:{
                                'drs_n': self.sheet[f'A{x}'].value,
                                'codigo_modelo': self.sheet[f'C{x}'].value,
                                'tipo_pedido_1':self.sheet[f'F{x}'].value,
                                'tipo_pedido_2':self.sheet[f'G{x}'].value,
                                'tipo_pedido_3':self.sheet[f'H{x}'].value,
                                'mercado':self.sheet[f'L{x}'].value,
                                'cliente':self.sheet[f'M{x}'].value,
                                'Uni_prod':self.sheet[f'S{x}'].value,
                                'Resp_pedido':self.sheet[f'U{x}'].value,
                                'dep_requerente':self.sheet[f'V{x}'].value,
                                'finalizado_por':self.sheet[f'AB{x}'].value,
                                'aprovado_por':self.sheet[f'AD{x}'].value,
                                'recusado_por':self.sheet[f'AE{x}'].value,
                                'anulado_por':self.sheet[f'AE{x}'].value,
                                'data_registo':str(self.sheet[f'AA{x}'].value),
                                'finalizado_em':str(self.sheet[f'AC{x}'].value)
                                }}
                    self.current_mounth_drs.update(new_dic)
            else:
                if recived_month == str(past_month):
                    if recived_year == str(self.year):
                    
                        new_dic = {drs_name:{
                                'drs_n': self.sheet[f'A{x}'].value,
                                'codigo_modelo': self.sheet[f'C{x}'].value,
                                'tipo_pedido_1':self.sheet[f'F{x}'].value,
                                'tipo_pedido_2':self.sheet[f'G{x}'].value,
                                'tipo_pedido_3':self.sheet[f'H{x}'].value,
                                'mercado':self.sheet[f'L{x}'].value,
                                'cliente':self.sheet[f'M{x}'].value,
                                'Uni_prod':self.sheet[f'S{x}'].value,
                                'Resp_pedido':self.sheet[f'U{x}'].value,
                                'dep_requerente':self.sheet[f'V{x}'].value,
                                'finalizado_por':self.sheet[f'AB{x}'].value,
                                'aprovado_por':self.sheet[f'AD{x}'].value,
                                'recusado_por':self.sheet[f'AE{x}'].value,
                                'anulado_por':self.sheet[f'AE{x}'].value,
                                'data_registo':str(self.sheet[f'AA{x}'].value),
                                'finalizado_em':str(self.sheet[f'AC{x}'].value)
                                }}
                        self.current_mounth_drs.update(new_dic)

            x += 1

        file.close()

   
    def getDrsTotalNum(self):
        return len(self.current_mounth_drs)
    

    def getDrsCanceled(self):
        data = self.current_mounth_drs
        drs_bucket = []
        for x in data:
            if self.current_mounth_drs[x]['anulado_por'] != None:
                drs_bucket.append(self.current_mounth_drs[x])
        return len(drs_bucket)
    

    def getAllDrsData(self):
        return self.current_mounth_drs


    def getAllDrsDataList(self):
        data = self.current_mounth_drs
        for x in data:
            print(x)
            print(self.current_mounth_drs[x])
            print('-'* 20)

    
    def GetNonAproved(self):
        data = self.current_mounth_drs
        drs_bucket = []
        for x in data:
            if self.current_mounth_drs[x]['aprovado_por'] == None and self.current_mounth_drs[x]['finalizado_por'] != None and self.current_mounth_drs[x]['anulado_por'] == None:
                drs_bucket.append(self.current_mounth_drs[x])
        return drs_bucket
    

    def GetAproved(self):
        data = self.current_mounth_drs
        drs_bucket = []
        for x in data:
            if self.current_mounth_drs[x]['aprovado_por'] != None:
                drs_bucket.append(self.current_mounth_drs[x])
        return drs_bucket
    


    def GetFinalized(self):
        data = self.current_mounth_drs
        drs_bucket = []
        for x in data:
            if self.current_mounth_drs[x]['finalizado_por'] == None:
                drs_bucket.append(self.current_mounth_drs[x])
        return drs_bucket


    def getMarkets(self):
        ''' Get all markets 
            return: DIC {tipo:qnt} '''

        data = self.current_mounth_drs

        tipos = []
        tipos_dic = {}
        for key, values in data.items():
            tipo = values['mercado']
            if tipo:
                tipos.append(tipo.upper().strip())
            
        for tip in tipos:
            y = {
                tip:0
            }
            tipos_dic.update(y)
        
        for x in tipos:
            tipos_dic[x] += 1


        dic_sorted = sorted(tipos_dic.items(), key=lambda x: x[1], reverse=True)
        return dic_sorted 


    def getClients(self):

        ''' Get all markets 
            return: DIC {tipo:qnt} '''

        data = self.current_mounth_drs

        tipos = []
        tipos_dic = {}
        for key, values in data.items():
            tipo = values['cliente']
            if tipo:
                tipos.append(tipo.upper().strip())
            
        for tip in tipos:
            y = {
                tip:0
            }
            tipos_dic.update(y)
        
        for x in tipos:
            tipos_dic[x] += 1


        dic_sorted = sorted(tipos_dic.items(), key=lambda x: x[1], reverse=True)
        return dic_sorted
    

    def getTypeRequest(self):
        ''' Get all markets 
            return: DIC {tipo:qnt} '''

        data = self.current_mounth_drs

        tipos = []
        tipos_dic = {}
        for key, values in data.items():
            tipo = values['tipo_pedido_1']
            if tipo != None:
                tipos.append(tipo)
            tipo_2 = values['tipo_pedido_2']
            if tipo_2 != None:
                tipos.append(tipo_2)
            tipo_3 = values['tipo_pedido_3']
            if tipo_3 != None:
                tipos.append(tipo_3)
            
        for tip in tipos:
            y = {
                tip:0
            }
            tipos_dic.update(y)
        
        for x in tipos:
            tipos_dic[x] += 1

        dic_sorted = sorted(tipos_dic.items(), key=lambda x: x[1], reverse=True)
        return dic_sorted
    

    def getRequestUnit(self):

        ''' Get all markets 
            return: DIC {tipo:qnt} '''

        data = self.current_mounth_drs

        tipos = []
        tipos_dic = {}
        for key, values in data.items():
            tipo = str(values['Uni_prod'])
            if tipo:
                tipos.append(tipo.upper().strip())
            
        for tip in tipos:
            y = {
                tip:0
            }
            tipos_dic.update(y)
        
        for x in tipos:
            tipos_dic[x] += 1


        dic_sorted = sorted(tipos_dic.items(), key=lambda x: x[1], reverse=True)
        return dic_sorted
    

    def getRequester(self):

        ''' Get all markets 
            return: DIC {tipo:qnt} '''

        data = self.current_mounth_drs

        tipos = []
        tipos_dic = {}
        for key, values in data.items():
            tipo = str(values['Resp_pedido'])
            if tipo:
                tipos.append(tipo.upper().strip())
            
        for tip in tipos:
            y = {
                tip:0
            }
            tipos_dic.update(y)
        
        for x in tipos:
            tipos_dic[x] += 1


        dic_sorted = sorted(tipos_dic.items(), key=lambda x: x[1], reverse=True)
        return dic_sorted
    

    def parseAllData(self):
        total_drs_n = self.getDrsTotalNum()
        non_aproved = self.GetNonAproved()
        non_aproved_n = len(non_aproved)
        finalized = self.GetFinalized()
        finalized_n = len(finalized)
        requests_types = self.getTypeRequest()
        clients = self.getClients()
        markets = self.getMarkets()
        productionUnit = self.getRequestUnit()
        requester = self.getRequester()
        canceled = self.getDrsCanceled()

        date = datetime.datetime.now()

        past_month = date.month - 1
        if past_month == 0:
            past_month = 12

        if self.global_ == True:
            past_month = 13
        #year = date.year

        data = {
            'total_drs_n': total_drs_n,
            'non_aproved': non_aproved,
            'non_aproved_n':non_aproved_n,
            'non_finalized':finalized,
            'non_finalized_n':finalized_n,
            'requests_types':requests_types,
            'clients':clients,
            'markets':markets,
            'ProductionUnit':productionUnit,
            'requester':requester,
            'canceled':canceled,
            'past_month':past_month,
            'year':self.year

        }

        return data
